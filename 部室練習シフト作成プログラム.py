import streamlit as st
import datetime
from datetime import date
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
import jpholiday
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.styles.alignment import Alignment
from mip import Model, xsum, minimize, BINARY, OptimizationStatus
from unicodedata import east_asian_width

#幅自動調整用データ
width_dict = {
  'F': 2,   # Fullwidth
  'H': 1,   # Halfwidth
  'W': 2,   # Wide
  'Na': 1,  # Narrow
  'A': 2,   # Ambiguous
  'N': 1    # Neutral
}
Font_depend = 1.2

#幅自動調整の関数
def sheet_adjusted_width(ws):
    # set column width
    for col in ws.columns:
        max_length= 1
        max_diameter = 1
        column= col[1].column_letter # Get the column name
        for cell in col:
            diameter = (cell.font.size*Font_depend)/10
            if diameter > max_diameter:
                max_diameter = diameter
            try:
                if(cell.value == None) : continue
                chars = [char for char in str(cell.value)]
                east_asian_width_list = [east_asian_width(char) for char in chars]
                width_list = [width_dict[east_asian_width] for east_asian_width in east_asian_width_list]
                if sum(width_list) > max_length:
                    max_length= sum(width_list)
            except:
                pass
            ws.column_dimensions[column].width= max_length*max_diameter + 1.2


tab_titles = ['部室練習固定シフト', '設営パートシフト']
tab1, tab2 = st.tabs(tab_titles)

border1 = Border(top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border2 = Border(top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_topthick = Border(top=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_bottomthick = Border(bottom=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_sidethick = Border(left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_topleft = Border(top=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
)

border_topcenter = Border(top=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border_topright = Border(top=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                bottom=Side(style='thin', color='000000')
)

border_bottomleft = Border(bottom=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border_bottomright = Border(bottom=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_bottomcenter = Border(bottom=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border_left = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border_right = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thick', color='000000')
)

border_allthin = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
)


band_list = {}
week = {}


st.session_state["page_control"] = 0
st.session_state["page_control2"] = 0
st.session_state["kinshi"] = {}

if "3pagenext" not in st.session_state:
  st.session_state["3pagenext"] = False
if "saitekika_button" not in st.session_state:
  st.session_state["saitekika_button"] = False


def wakusen(sheet):
  #枠線作成
  #B列目
  sheet.cell(row=2, column=2).border = border_topleft
  for t in range(1,8):
    sheet.cell(row=2+t, column=2).border = border_left
  sheet.cell(row=9, column=2).border = border_bottomleft
  #2行目
  for j in range(1, st.session_state["day_sum"]):
    sheet.cell(row=2, column=2+j).border = border_topcenter
  #右端
  sheet.cell(row=2, column=st.session_state["day_sum"]+2).border = border_topright
  for t in range(1,8):
    sheet.cell(row=2+t, column=st.session_state["day_sum"]+2).border = border_right
  sheet.cell(row=9, column=st.session_state["day_sum"]+2).border = border_bottomright
  #中
  for i in range(1, st.session_state["day_sum"]):
    for t in range(1,7):
      sheet.cell(row=2+t, column=2+i).border = border_allthin
  #下
  for j in range(1, st.session_state["day_sum"]):
    sheet.cell(row=9, column=2+j).border = border_bottomcenter

  #書式設定
  font = Font(name="游ゴシック",size=11,bold=True)
  for t in range(1, 9):
    for j in range(1, st.session_state["day_sum"]+2):
      sheet.cell(row=1+t, column=1+j).font = font
      sheet.cell(row=1+t, column=1+j).alignment = Alignment(horizontal = 'left', vertical = 'center')


def change_page():
  # ページ切り替えボタンコールバック
  st.session_state["page_control"] += 1

def change_page2():
  st.session_state["page_control2"] += 1

def band_list_making():
  i = 1
  while st.session_state["sheet"].cell(row=5 + i, column=2).value is not None:
      band_list[i] = st.session_state["sheet"].cell(row=5 + i, column=2).value
      i += 1
  band_sum = len(band_list)
  return band_sum

def option_select():
    st.session_state["max_practice"] = st.selectbox(
        '最大練習回数',
        [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
        index=0,
        placeholder="練習回数を選択してください"
    )
    return st.session_state["max_practice"]

def kinshi_select():
  # 日付を保存するためのセッション状態を初期化
  if 'dates_list' not in st.session_state:
    st.session_state['dates_list'] = []

  # 日付入力
  selected_date = st.date_input('日付を選択してください:', date.today())


  # 日付を追加するボタン
  if st.button('日付を追加'):
    if selected_date < st.session_state["start_day"] or selected_date > st.session_state["end_day"]:
      st.error('シフト期間内の日付を入力してください。')
      st.stop()
    st.session_state['dates_list'].append(selected_date)
    st.success(f'{selected_date} が追加されました。')

  if st.button('リセット'):
    st.session_state['dates_list'] = []
    st.success('リセットされました。')

  # 追加された日付と今日の日付の差分を辞書に登録（キーを1, 2, 3, ...とする）
  st.session_state["kinshi"] = {int(i + 1): (d - st.session_state["start_day"] + datetime.timedelta(days=1)).days for i, d in enumerate(st.session_state['dates_list'])}

  # 現在の追加済みの日付を表示
  st.write('追加された日付一覧:')
  i = 0
  while i < len(st.session_state['dates_list']):
    st.write(st.session_state['dates_list'][i])
    i += 1


  # # 日付の差分を含む辞書を表示
  # st.write('日付と今日との差分（日数）の辞書:', st.session_state["kinshi"])

def week_judge(start_day, vacation):
  # 平日と土日祝の判別
  current_day = start_day
  for i in range(1, st.session_state["day_sum"] + 1):
    if not vacation:
      if jpholiday.is_holiday(current_day) or current_day.weekday() in [5, 6]:
        week[i] = 1  # 祝日または土日
      else:
        week[i] = 0  # 平日
    else:
      week[i] = 1  # 長期休暇中
    current_day += datetime.timedelta(days=1)



def input_date():
  book1 = openpyxl.Workbook()
  for i in range(1, st.session_state["band_sum"] + 1):
    book1.create_sheet(index=0, title=band_list[i])
    sheet = book1[band_list[i]]
    for t in range(1, 8):
        sheet.cell(row=2 + t, column=2).value = str(t) + "限"
    calc_day = st.session_state["start_day"]
    j = 1
    while calc_day <= st.session_state["end_day"]:
        sheet.cell(row=2, column=2 + j).value = str(calc_day.month) + "/" + str(calc_day.day)
        j += 1
        calc_day += datetime.timedelta(days=1)

    wakusen(sheet)

    # デフォルトで作成されるシートを削除
  if 'Sheet' in book1.sheetnames:
      book1.remove(book1['Sheet'])

    # バイトストリームにExcelファイルを保存
  buffer = BytesIO()
  book1.save(buffer)
  buffer.seek(0)

  # StreamlitのダウンロードボタンでExcelファイルをダウンロード
  st.download_button(
      label="ダウンロード",
      data=buffer,
      file_name='シフト希望記入表.xlsx',
      mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
  )




def saitekika():
  # 最適化モデルの作成
  model = Model('PracticeShiftTime')
  # 決定変数の作成
  y = {(i, d, t): model.add_var(var_type='B') for i in st.session_state["I"] for d in st.session_state["D"] for t in st.session_state["T"]}
  st.session_state["y2"] = {(i, d, t): model.add_var(var_type='B') for i in st.session_state["I"] for d in st.session_state["D"] for t in st.session_state["T"]}
  s = {i: model.add_var(var_type='B') for i in st.session_state["I"]}

  # ハード制約の追加
  # 各バンドの制約
  for i in st.session_state["I"]:
      model += xsum(y[i, d, t] for d in st.session_state["D"] for t in st.session_state["T"]) >= 1  # 最低1回は練習
      model += xsum(y[i, d, t] for d in st.session_state["D"] for t in st.session_state["T"]) <= st.session_state["max_practice"]  # 練習回数は最大 st.session_state["max_practice"]
      for d in st.session_state["D"]:
          model += xsum(y[i, d, t] for t in st.session_state["T"]) <= 1  # 1日に1回のみ練習

  # 希望しない時間に練習を入れない
  for i in st.session_state["I"]:
      for d in st.session_state["D"]:
          for t in st.session_state["T"]:
              if st.session_state["kibou_time"][f"{i}_{d}_{t}"] == 0:
                  model += y[i, d, t] == 0

  
  # 部室利用禁止日に練習を割り当てない
  if st.session_state["kinshi"] is not None:
    for d in st.session_state["D"]:
        for k in st.session_state["kinshi"]:
            if d == st.session_state["kinshi"][k]:
              for i in st.session_state["I"]:
                for t in st.session_state["T"]:
                  model += y[i, d, t] == 0


  # 最終週に希望がある場合、必ず練習を入れる
  for i in st.session_state["I"]:
      if st.session_state["last_week"][i] == 1:
          model += xsum(y[i, d, t] for d in range(st.session_state["day_sum"] - 6, st.session_state["day_sum"] + 1) for t in st.session_state["T"]) >= 1 - s[i]

  # 同じ時間に練習するバンドは1つまで
  for d in st.session_state["D"]:
      for t in st.session_state["T"]:
          model += xsum(y[i, d, t] for i in st.session_state["I"]) <= 1

  # 連続して練習しない（1日以上あける）
  for i in st.session_state["I"]:
      for d in range(1, st.session_state["day_sum"]):
          model += xsum(y[i, d, t] for t in st.session_state["T"]) + xsum(y[i, d + 1, t] for t in st.session_state["T"]) <= 1

  # 目的関数の設定
  model.objective = minimize(-xsum(y[i, d, t] for i in st.session_state["I"] for d in st.session_state["D"] for t in st.session_state["T"]) + 10 * s[i])

  # 最適化の実行
  status = model.optimize()
  
  if status == OptimizationStatus.OPTIMAL:
    st.write('最適値 =', model.objective_value)
    st.session_state["y2"] = y
    for i in st.session_state["I"]:
      for d in st.session_state["D"]:
        for t in st.session_state["T"]:
          st.session_state["y2"][f"{i}_{d}_{t}"] = y[i, d, t].x

    result()


def result():
  print(st.session_state["y2"])
  book2 = openpyxl.Workbook()
  book2.create_sheet(index=0, title="結果出力")
  sheet = book2["結果出力"]

  for t in range(1, 8):
    sheet.cell(row=2 + t, column=2).value = str(t) + "限"
  calc_day = st.session_state["start_day"]
  j = 1
  while calc_day <= st.session_state["end_day"]:
      sheet.cell(row=2, column=2 + j).value = str(calc_day.month) + "/" + str(calc_day.day)
      j += 1
      calc_day += datetime.timedelta(days=1)
  
  for i in band_list:
    for d in range(1, st.session_state["day_sum"] + 1):
        for t in range(1, 8):
            if st.session_state["y2"][f"{i}_{d}_{t}"] > 0.01:
                sheet.cell(row=2 + t, column=2 + d).value = band_list[i]

  wakusen(sheet)

  # デフォルトで作成されるシートを削除
  if 'Sheet' in book2.sheetnames:
    book2.remove(book2['Sheet'])

  # sheet_adjusted_width(sheet)
  
  buffer2 = BytesIO()
  book2.save(buffer2)
  buffer2.seek(0)
  st.download_button(
    label="結果をダウンロード",
    data=buffer2,
    file_name="最適化結果.xlsx",
  mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



def practice_shift_main():
  # Webアプリのタイトル
  st.title('シフトスケジュール最適化')
  
  
  #ページ１：参加バンド登録
  uploaded_file_path = 'バンドリスト_テンプレート.xlsx'
  # ファイルをバイトとして読み込む
  with open(uploaded_file_path, 'rb') as file:
      band_listfile = file.read()
  
  if st.session_state["page_control"] == 0:
    st.header('１．参加バンドの登録')
    st.caption('ダウンロードボタンからテンプレートをダウンロードして、出演バンドを記入してください。')
    st.caption('記入を終えたファイルをアップロードしてください。')
  
    st.download_button(
        label="テンプレートをダウンロード",
        data=band_listfile,
        file_name='参加バンド登録＿テンプレート.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
  
    st.session_state["uploaded_file1"] = st.file_uploader("バンド名簿をアップロード", type=["xlsx"],key = "バンド名簿")
    if st.session_state["uploaded_file1"] is not None:
      change_page()
  
  
  
  #ページ２：ライブ情報の入力
  if "page_control" in st.session_state and st.session_state["page_control"] == 1:
      # st.session_state['uploaded'] = True
      st.header('２．ライブ情報の入力')
      st.session_state["book"] = load_workbook(st.session_state["uploaded_file1"])
      st.session_state["sheet"] = st.session_state["book"]["概要"]
      st.session_state["band_sum"] = band_list_making()
  
      st.session_state["start_day"] = st.date_input('シフト開始日を入力してください。', datetime.date(2024, 10, 10))
      st.session_state["end_day"] = st.date_input('シフト終了日を入力してください。', datetime.date(2024, 10, 31))
      if st.session_state["start_day"] > st.session_state["end_day"]:
          st.error('開始日は終了日より後の日付を入力してください。')
          st.stop()
  
      st.session_state["day_sum"] = (st.session_state["end_day"] - st.session_state["start_day"] + datetime.timedelta(days=1)).days
      st.session_state["max_practice"] = option_select()
      vacation = st.toggle("長期休暇期間")
      d = st.toggle("部室利用禁止日あり")
      if d == True:
        kinshi_select()
      week_judge(st.session_state["start_day"], vacation)
      
      if st.button("入力完了"):
        st.session_state["3pagenext"] = True
      if st.session_state["3pagenext"]:
        change_page()
  
  #ページ３：希望日入力
  if "page_control" in st.session_state and st.session_state["page_control"] == 2:
    st.header('３．練習希望日時の入力')
    input_date()
  
    #定数データ作成
    st.session_state["I"] = [i for i in range(1, st.session_state["band_sum"] + 1)]
    st.session_state["D"] = [i for i in range(1, st.session_state["day_sum"] + 1)]
    st.session_state["T"] = [i for i in range(1, 8)]
  
  
  
    st.write("記入を終えたファイルをアップロードしてください。")
  
    st.session_state["kibou_file"] = st.file_uploader(label="シフト希望表をアップロード", type=["xlsx"])
    
    if st.session_state["kibou_file"] is not None:
      change_page()
  
  if "page_control" in st.session_state and st.session_state["page_control"] == 3:
    st.header('４．最適化の実行')
    book1 = load_workbook(st.session_state["kibou_file"])
    st.session_state["kibou_time"] = {}
    st.session_state["last_week"] = {}
  
    #希望ファイルの読み込み
    for i in st.session_state["I"]:
        sheet_band = book1[band_list[i]]  # シートを取得
        for d in st.session_state["D"]:
            for t in st.session_state["T"]:
                value = sheet_band.cell(row=2 + t, column=2 + d).value
                if value is not 1:
                    value = 0      
                # キーを文字列に変換して保存
                key_str = f"{i}_{d}_{t}"
                st.session_state["kibou_time"][key_str] = value
  
    for i in st.session_state["I"]:
      st.session_state["last_week"][i] = int(any(st.session_state["kibou_time"][f"{i}_{d}_{t}"] for d in range(st.session_state["day_sum"] - 6, st.session_state["day_sum"] + 1) for t in st.session_state["T"]))
  
  
    if st.session_state["kibou_time"] is not None:
      st.write("希望の読み込みに成功しました。")
      st.write("実行ボタンを押してシフトを作成します。")
      if st.button("実行ボタン"):
        st.session_state["saitekika_button"] = True
      if st.session_state["saitekika_button"]:
        saitekika()
        
    else:
      st.write("希望の読み込みに失敗しました。もう一度ファイルを読み込ませてください。")

def part_shift_main():
  st.title('設営パートシフト最適化')
  
  #ページ１：参加バンド登録
  uploaded_file_path2 = 'パートシフト_テンプレート.xlsx'
  # ファイルをバイトとして読み込む
  with open(uploaded_file_path2, 'rb') as file:
    member_listfile = file.read()
  

    st.header('１．パートメンバーの登録')
    st.caption('ダウンロードボタンからテンプレートをダウンロードして、各パートに所属を記入してください。')
    st.caption('記入を終えたファイルをアップロードしてください。')
  
    st.download_button(
      label="テンプレートをダウンロード",
      data=member_listfile,
      file_name='パートメンバー登録＿テンプレート.xlsx',
      mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    st.session_state["member_file1"] = st.file_uploader("パート名簿をアップロード", type=["xlsx"],key = "パート名簿")
    if st.session_state["member_file1"] is not None:
      change_page2()

    #ページ２

    if "page_control2" in st.session_state and st.session_state["page_control2"] == 1:
      st.header('２．シフト希望の入力')
      st.caption('シフト希望入力表を作成しました。')
      st.caption('ダウンロードボタンからダウンロードし、記入を終えたファイルをアップロードしてください。')
      
      st.session_state["book"] = load_workbook(st.session_state["member_file1"])
      st.session_state["sheet"] = st.session_state["book"]["タイムテーブル"]
      Part_list=["ボーカル","ギター","ベース","ドラム","キーボード","PA","照明"]

      #タイムテーブルの読み込み
      st.session_state["timetable"]={}
      t=0
      while True:
        value = st.session_state["sheet"].cell(row=3+t, column=2).value
        if value is None:
          break
        st.session_state["timetable"][t] = value
        t += 1

      #インタミ直前のバンド数の読み込み
      intami = st.session_state["sheet"].cell(row=3, column=4).value


      book = openpyxl.Workbook()
      for Part in Part_list:
        #↓パートごとに分かれる
        st.session_state["sheet"] = st.session_state["book"][Part]

        i = 0  
        n1 = 0
        n2 = 0
        n3 = 0
      
        #3年生のデータを読み込む
        st.session_state["member"] = {}
        while True:
          value = st.session_state["sheet"].cell(row=4+i, column=2).value
          if value is None:
            break
          st.session_state["member"][value] =  st.session_state["sheet"].cell(row=4+i, column=3).value
          i += 1
          n3 += 1
      
        #2年生のデータを読み込む
        i = 0
        while True:
          value = st.session_state["sheet"].cell(row=4+i, column=5).value
          if value is None:
            break
          st.session_state["member"][value] =  st.session_state["sheet"].cell(row=4+i, column=6).value
          i += 1
          n2 += 1
      
        #1年生のデータを読み込む
        i = 0
        while True:
          value = st.session_state["sheet"].cell(row=4+i, column=8).value
          if value is None:
            break
          st.session_state["member"][value] =  st.session_state["sheet"].cell(row=4+i, column=9).value
          i += 1
          n1 += 1
        if n1  != 0 and n2 != 0 and n3 != 0:
          #希望用エクセルファイルの作成
          book.create_sheet(index=-1, title=Part)
          sheet = book[Part]
      
          #固定の文字列の作成
          sheet.cell(row=2, column=2).value = "総バンド数"
          sheet.cell(row=2, column=3).value = t
      
          sheet.cell(row=4, column=3).value = "人数"
          sheet.cell(row=5, column=2).value = "3年生"
          sheet.cell(row=6, column=2).value = "2年生"
          sheet.cell(row=7, column=2).value = "1年生"
      
          sheet.cell(row=5, column=3).value = n3
          sheet.cell(row=6, column=3).value = n2
          sheet.cell(row=7, column=3).value = n1
      
          sheet.cell(row=2, column=5).value = "インタミ直前"
          sheet.cell(row=2, column=6).value = intami
      
          #c[i,j,t]の表示
          sheet.cell(row=10, column=2).value = "c_(i,j,t)"
          sheet.cell(row=11, column=2).value = "3年生"
          sheet.cell(row=11+n3, column=2).value = "2年生"
          sheet.cell(row=11+n3+n2, column=2).value = "1年生"
      
          sheet.merge_cells(start_row=11, start_column=2, end_row=11+n3-1, end_column=2)
          sheet.merge_cells(start_row=11+n3, start_column=2, end_row=11+n2+n3-1, end_column=2)
          sheet.merge_cells(start_row=11+n2+n3, start_column=2, end_row=11+n1+n2+n3-1, end_column=2)
      
          sheet.merge_cells(start_row=10, start_column=3, end_row=10, end_column=4)
      
          #タイムテーブルの表示(横)
          for i in st.session_state["timetable"]:
            sheet.cell(row=10, column=5+i).value = st.session_state["timetable"][i]
      
          #パートメンバーの表示(縦)
          j=0
          for i in st.session_state["member"]:
            sheet.cell(row=11+j, column=3).value = str(i)
            sheet.cell(row=11+j, column=4).value = st.session_state["member"][i]
            j += 1
      
          #g_(i)を表示
          sheet.cell(row=10, column=4+t+2).value = "g_(i)"
          j = 1
          for i in st.session_state["member"]:
            if j >= n2+n3+1 and j < n1+n2+n3+1:
              sheet.cell(row=10+j-n2-n3, column=4+t+2).value = i
            j+=1

          #書式設定
          font = Font(name="游ゴシック",size=14,bold=True)
          for i in range(1,11+n1+n2+n3):
            for j in range(1,5+t+2):
              sheet.cell(row=1+i, column=1+j).font = font
              sheet.cell(row=1+i, column=1+j).alignment = Alignment(horizontal = 'left', vertical = 'center')

          #幅の自動調整(関数呼び出し)
          sheet_adjusted_width(sheet)
          
          #総バンド数の枠線
          sheet.cell(row=2, column=2).border = border1
          sheet.cell(row=2, column=3).border = border2

          #インタミ直前の枠線
          sheet.cell(row=2, column=5).border = border1
          sheet.cell(row=2, column=6).border = border2
        #各学年の部員の人数の枠線
    
          sheet.cell(row=4, column=2).border = border_topleft
          sheet.cell(row=4, column=3).border = border_topright
          sheet.cell(row=7, column=2).border = border_bottomleft
          sheet.cell(row=7, column=3).border = border_bottomright
          for i in range(5,7):
            sheet.cell(row=i, column=2).border = border_left
            sheet.cell(row=i, column=3).border = border_right
      
          #↓バンドによって出力が変わる
          #g[i]の枠線
          sheet.cell(row=10, column=5+t+1).border = border_topleft
          sheet.cell(row=10, column=5+t+2).border = border_topright
          sheet.cell(row=10+n1, column=5+t+1).border = border_bottomleft
          sheet.cell(row=10+n1, column=5+t+2).border = border_bottomright
          for i in range(11,10+n1):
            sheet.cell(row=i, column=5+t+1).border = border_left
            sheet.cell(row=i, column=5+t+2).border = border_right
      
          #c[i,t]の枠線
          #列Bの枠線
          sheet.cell(row=10, column=2).border = border_topthick
      
          sheet.cell(row=11, column=2).border = border_topthick
          sheet.cell(row=11+n3, column=2).border = border_topthick
          sheet.cell(row=11+n3+n2, column=2).border = border_topthick
      
          sheet.cell(row=11+n3-1, column=2).border = border_bottomthick
          sheet.cell(row=11+n2+n3-1, column=2).border = border_bottomthick
          sheet.cell(row=11+n3+n2+n1-1, column=2).border = border_bottomthick
          for i in range(12,12+n3-2):
              sheet.cell(row=i, column=2).border = border_sidethick
          for i in range(12+n3,12+n3+n2-2):
              sheet.cell(row=i, column=2).border = border_sidethick
          for i in range(12+n3+n2,12+n3+n2+n1-2):
              sheet.cell(row=i, column=2).border = border_sidethick
      
          #列C,Dの枠線
          sheet.cell(row=10, column=3).border = border_topleft
          sheet.cell(row=10, column=4).border = border_topright
          #3回生
          sheet.cell(row=11+n3-1, column=3).border = border_bottomleft
          sheet.cell(row=11+n3-1, column=4).border = border_bottomright
          sheet.cell(row=11, column=3).border = border_topleft
          sheet.cell(row=11, column=4).border = border_topright
          for i in range(12,10+n3):
            sheet.cell(row=i, column=3).border = border_left
            sheet.cell(row=i, column=4).border = border_right
      
          #2回生
          sheet.cell(row=11+n2+n3-1, column=3).border = border_bottomleft
          sheet.cell(row=11+n2+n3-1, column=4).border = border_bottomright
          sheet.cell(row=11+n3, column=3).border = border_topleft
          sheet.cell(row=11+n3, column=4).border = border_topright
          for i in range(12+n3,10+n2+n3):
            sheet.cell(row=i, column=3).border = border_left
            sheet.cell(row=i, column=4).border = border_right
      
          #1回生
          sheet.cell(row=11+n1+n2+n3-1, column=3).border = border_bottomleft
          sheet.cell(row=11+n1+n2+n3-1, column=4).border = border_bottomright
          sheet.cell(row=11+n2+n3, column=3).border = border_topleft
          sheet.cell(row=11+n2+n3, column=4).border = border_topright
          for i in range(12+n2+n3,10+n1+n2+n3):
            sheet.cell(row=i, column=3).border = border_left
            sheet.cell(row=i, column=4).border = border_right
      
          #列E～の枠線
          sheet.cell(row=10, column=5).border = border_topleft
          sheet.cell(row=11, column=5).border = border_topleft
          sheet.cell(row=10+n1+n2+n3, column=5).border = border_bottomleft
      
          for i in range(6, 6+t-2):
            sheet.cell(row=10, column=i).border = border_topcenter
            sheet.cell(row=11, column=i).border = border_topcenter
            sheet.cell(row=10+n1+n2+n3, column=i).border = border_bottomcenter
      
      
          #右端の枠線
          sheet.cell(row=10, column=6+t-2).border = border_topright
          sheet.cell(row=11, column=6+t-2).border = border_topright
          sheet.cell(row=10+n1+n2+n3, column=6+t-2).border = border_bottomright
      
          #右,左端の枠線
          for i in range(12,12+n1+n2+n3-2):
            sheet.cell(row=i, column=5).border = border_left
            sheet.cell(row=i, column=6+t-2).border = border_right
      
          for i in range(12,12+n1+n2+n3-2):
            for j in range(6,6+t-2):
              sheet.cell(row=i, column=j).border = border_allthin
      
      book.remove(book['Sheet'])

     # バイトストリームにExcelファイルを保存
      buffer = BytesIO()
      book.save(buffer)
      buffer.seek(0)
    
      # StreamlitのダウンロードボタンでExcelファイルをダウンロード
      st.download_button(
          label="ダウンロード",
          data=buffer,
          file_name='シフト希望記入表.xlsx',
          mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
      )
      st.session_state["kibou_file2"] = st.file_uploader("シフト希望をアップロード", type=["xlsx"])
      if st.session_state["kibou_file2"] is not None:
        change_page2()
        
    if "page_control2" in st.session_state and st.session_state["page_control2"] == 2:
      st.header('３．最適化の実行')
      st.caption('シフト希望入力表の読み込みが完了しました。')
      st.caption('実行ボタンを押して最適化を実行してください。')
      
      if st.button("実行ボタン"):
        st.session_state["saitekika_button"] = True
      if st.session_state["saitekika_button"]:
        change_page2()
        
    if "page_control2" in st.session_state and st.session_state["page_control2"] == 3:
      #出力用ファイルの作成
      for Part in Part_list:
        if Part == "ボーカル":
          book = openpyxl.Workbook()
        else:
          book = load_workbook('2023年学祭パートシフト表(外ステージ).xlsx')

        #パートシートの追加
        book.create_sheet(index=-1, title=Part)
        sheet = book[Part]
    
        #インタミ要素の追加
        T=[]
        tt = 0
        for i in st.session_state["timetable"]:
          T.append(st.session_state["timetable"][i])
          tt += 1
        T.insert(intami,"インタミ")
        tt += 1
        timetable_new = {}
        for i in range(1,tt+1):
          timetable_new[i] = T[i-1]
    
        m=7
        n=1

        
        #c[i,j,t]の表示
        sheet.cell(row=3, column=2).value = Part
        sheet.cell(row=11-m, column=2+n).value = "3年生"
        sheet.cell(row=11+n3-m, column=2+n).value = "2年生"
        sheet.cell(row=11+n3+n2-m, column=2+n).value = "1年生"
    
        sheet.merge_cells(start_row=3, start_column=2, end_row=3+n1+n2+n3, end_column=2)
    
        sheet.merge_cells(start_row=11-m, start_column=2+n, end_row=11+n3-1-m, end_column=2+n)
        sheet.merge_cells(start_row=11+n3-m, start_column=2+n, end_row=11+n2+n3-1-m, end_column=2+n)
        sheet.merge_cells(start_row=11+n2+n3-m, start_column=2+n, end_row=11+n1+n2+n3-1-m, end_column=2+n)
    
        sheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
    
        #タイムテーブルの表示(横)
        for i in timetable_new:
          sheet.cell(row=10-m, column=5+i).value = timetable_new[i]
    
        #パートメンバーの表示(縦)
        j=0
        for i in st.session_state["member"]:
          sheet.cell(row=11+j-m, column=3+n).value = i
          sheet.cell(row=11+j-m, column=4+n).value = st.session_state["member"][i]
          j += 1
    
        #書式設定
        font = Font(name="游ゴシック",size=14,bold=True)
        for i in range(1,30):
          for j in range(1,30):
            sheet.cell(row=1+i, column=1+j).font = font
            sheet.cell(row=1+i, column=1+j).alignment = Alignment(horizontal = 'left', vertical = 'center')
    
    
        #幅の自動調整(関数呼び出し)
        sheet_adjusted_width(sheet)
    
        #列Bの枠線
        sheet.cell(row=3, column=2).border = border_topthick
        sheet.cell(row=3+n1+n2+n3, column=2).border = border_bottomthick
        for i in range(4,3+n1+n2+n3):
          sheet.cell(row=i, column=2).border = border_sidethick
    
    
        #列Cの枠線
        sheet.cell(row=3, column=3).border = border_topleft
    
        sheet.cell(row=4, column=3).border = border_topthick
        sheet.cell(row=4+n3, column=3).border = border_topthick
        sheet.cell(row=4+n3+n2, column=3).border = border_topthick
    
        sheet.cell(row=4+n3-1, column=3).border = border_bottomthick
        sheet.cell(row=4+n2+n3-1, column=3).border = border_bottomthick
        sheet.cell(row=4+n3+n2+n1-1, column=3).border = border_bottomthick
        for i in range(5,5+n3-2):
            sheet.cell(row=i, column=3).border = border_sidethick
        for i in range(5+n3,5+n3+n2-2):
            sheet.cell(row=i, column=3).border = border_sidethick
        for i in range(5+n3+n2,5+n3+n2+n1-2):
            sheet.cell(row=i, column=3).border = border_sidethick
    
        #列D,Eの枠線
        #3回生
        sheet.cell(row=3, column=4).border = border_topcenter
        sheet.cell(row=3, column=5).border = border_topright
    
        sheet.cell(row=4+n3-1, column=4).border = border_bottomleft
        sheet.cell(row=4+n3-1, column=5).border = border_bottomright
        for i in range(5,5+n3-2):
          sheet.cell(row=i, column=4).border = border_left
          sheet.cell(row=i, column=5).border = border_right
        sheet.cell(row=4, column=4).border = border_topleft
        sheet.cell(row=4, column=5).border = border_topcenter
        sheet.cell(row=4, column=6).border = border_topright
    
    
    
        #2回生
        for i in range(5+n3,5+n2+n3-2):
          sheet.cell(row=i, column=4).border = border_left
          sheet.cell(row=i, column=5).border = border_right
        sheet.cell(row=4+n2+n3-1, column=4).border = border_bottomleft
        sheet.cell(row=4+n2+n3-1, column=5).border = border_bottomright
        sheet.cell(row=4+n3, column=4).border = border_topleft
        sheet.cell(row=4+n3, column=5).border = border_topright
    
    
        #1回生
        for i in range(5+n2+n3,5+n1+n2+n3-2):
          sheet.cell(row=i, column=4).border = border_left
          sheet.cell(row=i, column=5).border = border_right
        sheet.cell(row=4+n1+n2+n3-1, column=4).border = border_bottomleft
        sheet.cell(row=4+n1+n2+n3-1, column=5).border = border_bottomright
        sheet.cell(row=4+n2+n3, column=4).border = border_topleft
        sheet.cell(row=4+n2+n3, column=5).border = border_topright
    
    
    
        #列E～の枠線
        sheet.cell(row=3, column=6).border = border_topleft
        sheet.cell(row=4, column=6).border = border_topleft
        sheet.cell(row=3+n1+n2+n3, column=6).border = border_bottomleft
    
        for i in range(7, 7+tt-2):
          sheet.cell(row=3, column=i).border = border_topcenter
          sheet.cell(row=4, column=i).border = border_topcenter
          sheet.cell(row=3+n1+n2+n3, column=i).border = border_bottomcenter
    
        #右端の枠線
        sheet.cell(row=3, column=7+tt-2).border = border_topright
        sheet.cell(row=4, column=7+tt-2).border = border_topright
        sheet.cell(row=3+n1+n2+n3, column=7+tt-2).border = border_bottomright
    
        #右,左端の枠線
        for i in range(5,5+n1+n2+n3-2):
          sheet.cell(row=i, column=6).border = border_left
          sheet.cell(row=i, column=7+tt-2).border = border_right
    
        for i in range(5,5+n1+n2+n3-2):
          for j in range(7,7+tt-2):
            sheet.cell(row=i, column=j).border = border_allthin
    
        #インタミ塗りつぶし
        fill = PatternFill(patternType='solid', fgColor='d3d3d3')
        for i in range(3,3+n1+n2+n3+1):
          sheet.cell(row=i, column=6+intami).fill = fill
    
        if Part == "照明":
          book.remove(book['Sheet'])
    
       # バイトストリームにExcelファイルを保存
        buffer = BytesIO()
        book.save(buffer)
        buffer.seek(0)
      
        # StreamlitのダウンロードボタンでExcelファイルをダウンロード
        st.download_button(
            label="ダウンロード",
            data=buffer,
            file_name='test.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


      
    #   st.session_state["book"] = load_workbook(st.session_state["kibou_file2"])
      # for Part in Part_list:
      #   st.session_state["sheet"] = st.session_state["book"][Part]
      #   jouken = [0,0,0,0,0]
      #   if Part == "ボーカル":
      #     jouken=[3,0,1,0,1,1]
      #   elif Part == "ギター":
      #     jouken=[3,0,1,0,1,1]
      #   elif Part == "ベース":
      #     jouken=[3,0,1,0,1,1]
      #   elif Part == "PA":
      #     jouken=[4,1,1,0,1,1]
      #   elif Part == "照明":
      #     jouken=[4,1,1,0,1,1]
      #   elif Part == "ドラム":
      #     jouken=[4,1,1,1,0,0]


      #   #総バンド数
      #   m = sheet.cell(row=2, column=3).value
      
      #   #各学年の人数
      #   n1 = sheet.cell(row=7, column=3).value
      #   n2 = sheet.cell(row=6, column=3).value
      #   n3 = sheet.cell(row=5, column=3).value
      
      
      #   #定数用のデータの作成
      #   I = [i+1 for i in range(n1+n2+n3)]
      #   T = [i+1 for i in range(m)]
      
      #   #インタミ直前のバンド
      #   intami = sheet.cell(row=2, column=6).value
      
      #   c ={} #出演都合
      #   for i in I:
      #     for t in T:
      #       value = sheet.cell(row=10+i, column=4+t).value
      #       c[i, t] = value if value is not None else 1
      
      #   g = {} #1回生の講習会参加
      #   for i in range(n2+n3+1,n1+n2+n3+1):
      #     value = sheet.cell(row=10+i-n2-n3, column=18).value
      #     g[i] = value if value is not None else 0
      
      #   #空問題の作成
      #   model = Model('PartShift')


      #   #決定変数の作成
      #   x = {}
      #   for i in I:
      #     for t in T:
      #       x[i, t] = model.add_var(f'x{i},{t}', var_type='B')
      
      #   y = {}
      #   for i in I:
      #     for j in I:
      #       for t in T:
      #         y[i,j,t] = model.add_var(f'y{i},{j},{t}', var_type='B')
      
      #   z = {}
      #   for i in I:
      #     for j in I:
      #       z[i,j] = model.add_var(f'z{i},{j}', var_type='B')
      
      #   #ペナルティ変数
      #   w = {}
      #   for i in I:
      #     for t in T:
      #       w[i, t] = model.add_var(f'w{i},{t}', var_type='B')
      
      #   v = {}
      #   for i in I:
      #     for t in range(1,m):
      #       v[i, t] = model.add_var(f'v{i},{t}', var_type='B')
      
      #   u = {}
      #   for i in I:
      #     for t in range(1,m):
      #       u[i, t] = model.add_var(f'u{i},{t}', var_type='B')
      
      #   s = {}
      #   for i in I:
      #     for j in I:
      #       s[i,j] = model.add_var(f's{i},{j}', var_type='B')
      
      
      #   #制約条件の追加
      
      
      #   #①出演直後のメンバーに仕事を割り当てない
      #   #②講習会に参加していない1回生が仕事を割り当てられた場合、必ず2回生が２人以上同じスロットに入る
      #   #③できるだけメンバーは連続して仕事をしない
      #   #④できるだけ出演前のシフトに仕事を割り当てない
      #   #⑤なるべく違う部員と仕事をする
      
      #   #ハード制約条件
      #   #スロットｔには2人以上,4人以下を割り当てる
      #   for t in T:
      #     model += xsum(x[i,t] for i in I) >= 2
      #     model += xsum(x[i,t] for i in I) <= jouken[0]
      
      #   #連続して仕事ができるのは最大3回まで
      #   if jouken[3] == 0:
      #     for i in I:
      #       for t in range(1,m-2):
      #         model += x[i,t] + x[i,t+1] + x[i,t+2] +x[i,t+3] <= 3
      
      #   #希望スロット以外に仕事を割り振らない
      #   for i in I:
      #     for t in T:
      #       if c[i,t] == 2 or c[i,t] == 0:
      #         model += x[i,t] == 0
      #       # elif c[i,t] == 1:
      #       #   model += x[i,t] <= 1
      
      #   #1回生が仕事を割り当てられた場合、必ず2回生が1人以上同じスロットに入る
      #   for t in T:
      #     model += xsum(x[i,t] for i in range(n2+n3+1, n1+n2+n3+1)) <= xsum(x[i,t] for i in range(1,n2+n3+1))*3
      
      #   #講習会に参加していない1回生は最低2回以上仕事をする
      #   for i in range(n2+n3+1,n1+n2+n3+1):
      #     if xsum(c[i,t] for t in T) >= 1:
      #       model += xsum(x[i,t] for t in T) >= 1
      
      #   #①出演直後のメンバーに仕事を割り当てない
      #   if jouken[1] == 1:
      #     for i in I:
      #       for t in range(1,m):
      #         if c[i,t] == 2:
      #           if t != intami:
      #             model += x[i,t+1] == 0
      
      
      #   #ソフト制約条件
      #   #②講習会に参加していない1回生が仕事を割り当てられた場合、必ず2回生が２人以上同じスロットに入る
      #   for i in range(n2+n3+1, n1+n2+n3+1):
      #     if g[i] == 1:
      #       for t in T:
      #         model += xsum(x[j,t] for j in range(1,n2+n3+1)) >= 2 * x[i,t] - w[i,t]
      
      #   #③できるだけメンバーは連続して仕事をしない
      #   for i in I:
      #     for t in range(1,m):
      #       model += x[i,t] + x[i,t+1] <= 1 + v[i,t]
      
      #   #④できるだけ出演前のシフトに仕事を割り当てない
      #   for i in I:
      #     for t in range(1,m):
      #       if c[i,t+1] == 2:
      #         model += x[i,t] <= u[i,t]
      
      #   #⑤なるべく違う部員と仕事をする
      #   if n2 < n1:
      #     for i in range(n2+1,n2+n3+1):
      #       for j in I:
      #         for t in T:
      #           if i != j:
      #             model += x[i,t] + x[j,t] >= 2*y[i,j,t]
      
      #     for i in range(n2+1,n2+n3+1):
      #       for j in I:
      #         if i != j:
      #           model += xsum(y[i,j,t] for t in T) >= z[i,j] - s[i,j]
      #   elif n1 <= n2:
      #     for i in range(n2+n3+1,n1+n2+n3+1):
      #       for j in I:
      #         for t in T:
      #           if i != j:
      #             model += x[i,t] + x[j,t] >= 2*y[i,j,t]
      
      #     for i in range(n2+n3+1,n1+n2+n3+1):
      #       for j in I:
      #         if i != j:
      #           model += xsum(y[i,j,t] for t in T) >= z[i,j] - s[i,j]
      
      #   #最適化
      #   #目的関数の設定
      #   model.objective = minimize(jouken[2]*xsum(w[i,t] for i in I for t in T) + jouken[3]*xsum(v[i,t] for i in I for t in range(1,m)) + jouken[4]*xsum(u[i,t] for i in I for t in range(1,m))
      #   -jouken[5]*xsum(z[i,j] for i in I for j in I)
      #   + jouken[5]*1.1*xsum(s[i,j] for i in I for j in I)
      #   +xsum(x[i,t] for i in range(n3+1,n2+n3+1) for t in T) +5*xsum(x[i,t] for i in range(1,n3+1) for t in T))
      
      #   #最適化の実行
      #   status = model.optimize()

    
        
  

with tab1:
  practice_shift_main()

with tab2:
  part_shift_main()
